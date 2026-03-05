import numpy as np
import CC_main_V2 as CC
from openpyxl import load_workbook
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import os

def check_date(parent_folder,subfolder_to_check):
    """This function checks if the subfolder to check is the latest subfolder in the parent folder. It returns True if it is the latest, and False if it is not.
    Input:
    parent_folder: (String) The path to the parent folder where the subfolders are located.
    subfolder_to_check: (String) The name of the subfolder to check if it is the latest.
    Output:
    (Boolean) True if the subfolder to check is the latest, False if it is not."""

    latest_subfolder = None
    latest_mtime = -1

    with os.scandir(parent_folder) as entries:
        for entry in entries:
            if entry.is_dir(follow_symlinks=False):
                if entry.stat().st_mtime > latest_mtime:
                    latest_mtime = entry.stat().st_mtime
                    latest_subfolder = entry.name

    return subfolder_to_check == latest_subfolder


def extract_thermokoppels_meetleidingen(file_path):
    """"
    This function reads in the intial information of the curecycle from the csv file. It provides some key data and the temperatures and pressures per bag in the cure.
    Input:
    file_path: (String) The path to the csv file containing the curecycle data.
    Output:
    meetleidingen_per_zak: (List of lists) A list containing the unique meetleidingen per zak.
    header_row: (Integer) The row number where the header of the data starts.
    CureCycle: (String) The curecycle which is being analysed (e.g. CC0092).
    WO_lst: (List) A list containing the WO numbers which are included in this curecycle.
    thermokoppels_per_zak: (List of lists) A list containing the unique thermokoppels per WO.
    """
    WO_lst = []

    with open(file_path, 'r', encoding='utf-8') as f:
        for i, line in enumerate(f):
            line = line.strip()

            if line == "Producten:":
                row_start = i
            elif line == "Logging:":
                row_end = i-1
            elif 'CurveNr' in line:
                CureCycle = line.split(',')[1].split(' ')[0]

            elif "Timestamp" in line:  # Replace with something from your column names
                header_row = i+1
                break

    df = pd.read_csv(file_path,delimiter=',',header=row_start,nrows=int(row_end-row_start-2)) #Index(['Artikelnummer', 'POnumber', 'MalNr', 'MalID', 'ZakNummer','PDrukAlgoritmeID', 'Thermokoppels', 'Instelleidingen', 'Meetleidingen','TkTags', 'IlTags', 'MlTags'], dtype='object')


    meetleidingen_per_zak = []
    thermokoppels_per_zak = []

    # Group by ZakNummer and process
    grouped_meet = df.groupby('POnumber')['Meetleidingen']
    grouped_tk = df.groupby('POnumber')['Thermokoppels']

    for WO, group in grouped_meet:
        # Split strings into individual codes and flatten the list
        codes = set()
        for entry in group:
            codes.update(entry.split('.'))

        # Store result as tuple (ZakNummer, [unique codes])
        meetleidingen_per_zak.append( sorted(codes))
        WO_lst.append(WO)
    
    for WO, group in grouped_tk:
        # Split strings into individual codes and flatten the list
        codes = set()
        for entry in group:
            codes.update(entry.split('.'))

        # Store result as tuple (ZakNummer, [unique codes])
        thermokoppels_per_zak.append( sorted(codes))

    return meetleidingen_per_zak,header_row,CureCycle,WO_lst,thermokoppels_per_zak
     

def read_in_CCxx(fname):
    """
    This function takes in a filename which refers to the requirements of a specific cure. It will read in all relevant data and output the requirements in lists and variables.

    Input:
    fname: (String) Should refer to the filename which corresponds to the 

    Output: all important requirements for the curecycle in the form of lists and variables

    """

    workbook = load_workbook(fname, data_only=True)  # data_only=True avoids formulas
    sheet = workbook['Sheet1']  # Replace with your actual sheet name

    # === Step 1: Read the header row ===
    cycles = False
    rates = False
    phases_lst = []
    DT_lst = []
    Dt_lst = []
    Heat_up_rate = []
    for row in sheet.iter_rows(min_row=1, max_row=100):  # search first 10 rows for headers

        headers = [cell.value for cell in row]
        if any(headers):  # only use non-empty rows
            header_row = headers
            #Read in initial conditions
            if header_row[0] == "Cool down rate":
                Cool_down_rate = header_row[2], header_row[3]
            if header_row[0] == "Autoclave Pressure range":
                AP = [header_row[2], header_row[3]]
            if header_row[0] == "Vacuum bag pressure range":
                vac_range = [header_row[2], header_row[3]]

            if header_row[0] == "Begin Temperature":
                Start_cure_cond = [header_row[2], None]
            if header_row[0] == "End Temperature":
                End_cure_cond = [header_row[2], None]
            if header_row[0] == "Begin Time":
                Start_cure_cond[1] = header_row[2]
            # if header_row[0] == "End Time":
            #     End_cure_cond[1] = header_row[2]
            if header_row[0] == "Leak test":
                deltaP_thresh = header_row[2]
            if header_row[0] == "Allowed temperature spread":
                Spread_limit = header_row[2]

            # read in cures
            if header_row[0] == 'Cure Cycle Stages':
                cycles = True
            if header_row[0] == 'Heat up cycles': # introduced for if different heat-up rates will be used
                cycles = False
                rates = True

            # define order of phases and store
            if cycles and header_row[0] is not None:
                if header_row[0].lower() in ['heat-up', 'cure', 'cooldown', 'dwell']:

                    if header_row[0].lower() == 'dwell':
                        phases_lst.append('heatup')
                        phases_lst.append(header_row[0].lower())
                    elif header_row[0].lower() == 'cure':
                        phases_lst.append('heatup')
                        phases_lst.append(header_row[0].lower())
                        phases_lst.append('cooldown')
                    else:
                        phases_lst.append(header_row[0].lower())

                    DT_lst.append([header_row[1], header_row[2]]) # List of min and max dwell temperatures [degC] Cure is seen as dwell
                    Dt_lst.append([header_row[3], header_row[4]]) # List of min and max dwell time [min] Cure is seen as dwell

            if rates and header_row[0] is not None:
                if header_row[0] == "Heat-up":
                    Heat_up_rate.append([header_row[1], header_row[2]])
                

    return Heat_up_rate, Cool_down_rate, AP, vac_range, DT_lst, Dt_lst, deltaP_thresh, Start_cure_cond, End_cure_cond, phases_lst, Spread_limit

def temperatures(df,thermokoppels):
    """"
    From the provided temperature data this function extracts the leading and lagging TC temperatures. It will also check if at any given moment at least 3 TCs are recording if not, a warning is provided."""
    row_max_values = []
    row_min_values = []
    Warning_amount_TCs = False
    for _, row in df.iterrows():
        valid_values = []

        for tc in thermokoppels:
            val = row[f"Thermokoppel {tc[-2::]} Waarde"] 
            status = row[f"Thermokoppel {tc[-2::]} Statusnummer"]
            

            if status == 1:
                valid_values.append(val)

        # If no TC has status 0 → store NaN
        if len(valid_values)>0:
            row_max_values.append(max(valid_values))
            row_min_values.append(min(valid_values))
            if len(valid_values)<3:
                Warning_amount_TCs = True
        else:
            Warning_amount_TCs = True
            row_max_values.append(np.nan)
            row_min_values.append(np.nan)

    return np.array(row_max_values), np.array(row_min_values), Warning_amount_TCs


##This function performs the calculation once per curecycle to quickly change between WOs without having to recalculate everything.
@st.cache_data(show_spinner=True) #This line ensures that the curecycle data is only loaded once and that the UI can just show new pages without a compllete recalcualtion/
def analyse_curecycle(cc_num):
    """
    This is the main function and reads in the data + requirementts to provide a log. Here all necessary information is collected and calculated. Later one it
    different pieces of this information will be shown.
    :param cc_num:
    :return:
    """

    Unit_csv = "sec" # Might still be needed in certain function but should remain set on seconds as sometimes this is alreaady hardcoded and sometimes its a choice.
    # fname = f"\\\\air-dc1-fil010\\PLC_data\\autoclave\\ProofOfProcess_IACT\\LogData\\Charge Reports\\Charge {cc_num}\\Charge {cc_num} DE02 Logging seconden.csv"
    fname = f'C:\\Users\\oscar\\PycharmProjects\\CureCycle_Checker\\CureCycle Data\\Charge {cc_num} DE02 Logging seconden.csv'
    # is_latest = check_date(f"\\\\air-dc1-fil010\\PLC_data\\autoclave\\ProofOfProcess_IACT\\LogData\\Charge Reports", f"\\\\air-dc1-fil010\\PLC_data\\autoclave\\ProofOfProcess_IACT\\LogData\\Charge Reports\\Charge {cc_num}")

    is_latest = check_date(f"C:\\Users\\oscar\\PycharmProjects\\CureCycle_Checker\\CureCycle Data",f"C:\\Users\\oscar\\PycharmProjects\\CureCycle_Checker\\CureCycle Data\\Charge {cc_num}")
    
    meetleidingen_per_zak,header_row,CureCycle,WO_lst,thermokoppels_per_zak = extract_thermokoppels_meetleidingen(fname)

    #It will search if any requirements document is available for this curecycle, if not it will provide a warning and a different curecycle should be chosen.
    try:
        # fname2 = f'C:\\Users\\Oscar.Carpentier\\OneDrive - Airborne\\Documents\\Visual Studios Code\\Check_CureCycles\\{CureCycle}.xlsx'
        fname2 = f'C:\\Users\\oscar\\PycharmProjects\\CureCycle_Checker\\CureCycle Templates\\{CureCycle}.xlsx'
        Heat_up_rate, Cool_down_rate, AP, vac_range, DT_lst, Dt_lst, deltaP_thresh, Start_cure_cond, End_cure_cond, phases_lst, Spread_limit = read_in_CCxx(fname2)
        
    except: 
        st.warning(f'Curecycle {CureCycle} has no approved requirements document')

    #read in all important data and put in arrays for later use:
    df = pd.read_csv(fname,skiprows=header_row) # could differ by a line or so
    T_air = df[['AC Temp 01 Waarde','AC Temp 02 Waarde','AC Temp 03 Waarde']].max(axis=1).to_numpy()
    
    Press = df['AC druk Waarde'].to_numpy()
    if Unit_csv == 'sec':
        time = df['Seconden'].to_numpy()
    elif Unit_csv == 'min':
        time = df['Minuten'].to_numpy()

    
    #define list to store data:
    leaktest_log = []
    Vacuum_log = []
    cure_superlog = []
    pressure_super_log = []
    spread_super_log = []
    T_lead_lst = []
    T_lag_lst = []
    Warning_TC_lst = []
    Pressure_meas_lst = []
    Pressure_drive_lst = []
    nok_found_lst = []
    for idx,WO in enumerate(WO_lst):
        T_lead,T_lag,Warning_TCs = temperatures(df,thermokoppels_per_zak[idx])
        T_lead_lst.append(T_lead)
        T_lag_lst.append(T_lag)
        Warning_TC_lst.append(Warning_TCs)

        # Divide the different cycles to evaluate them separately
        
        cure_superlog.append(CC.Check_cycles(time, T_lead, T_lag, T_air, phases_lst, Start_cure_cond, End_cure_cond, DT_lst, Dt_lst,Heat_up_rate, Cooldownrate = Cool_down_rate,Unit_csv=Unit_csv))
        pressure_super_log.append(CC.Check_Pressure(time, Press, AP, T_lag, T_lead, Start_cure_cond, End_cure_cond,Unit_csv=Unit_csv))
        spread_super_log.append(CC.spread(T_lag,T_lead,time,Spread_limit))

        Pressure_meas_lst.append([])
        Pressure_drive_lst.append([])
        meetleidingen = meetleidingen_per_zak[idx]

        Vacuum_log.append([])
        leaktest_log.append([])
        for num in meetleidingen:
            num = num.strip('ML')
            
            Pressure_meas, Presssure_drive = df[f'Meetleiding {num} Waarde'].to_numpy(),df[f'Meetleiding {num}'].to_numpy()
            Pressure_meas_lst[-1].append(Pressure_meas)
            Pressure_drive_lst[-1].append(Presssure_drive)
            leaktest_log[-1].append([num,CC.check_leaktest(time, Pressure_meas, Presssure_drive, deltaP_thresh)])
            Vacuum_log[-1].append([num,CC.Check_Vacuum(time, Pressure_meas, Presssure_drive, vac_range, T_lag, T_lead, Start_cure_cond, End_cure_cond, Unit_csv)])

        #This line will check for each WO if there is an error. If there is anything NOK then the boolean will be False.
        nok_found = any(log[0] == 1 for log in cure_superlog[idx]) or \
                    any(log[1][0] == 1 for log in leaktest_log[idx]) or \
                    any(log[1][0][0] == 1 for log in Vacuum_log[idx]) or \
                    any(log[0][0]==1 for log in pressure_super_log) 
        
        nok_found_lst.append(nok_found)
    return time, T_lead_lst, T_lag_lst, T_air, Press, Warning_TC_lst, cure_superlog, pressure_super_log, spread_super_log, leaktest_log, Vacuum_log, meetleidingen_per_zak, header_row,CureCycle, WO_lst, thermokoppels_per_zak, AP, vac_range, Pressure_meas_lst,  Pressure_drive_lst, nok_found_lst, is_latest


def plot_temp(time,T_lag,T_lead,T_air):
    df_plot = pd.DataFrame({
        "Time [min]": time / 60,
        "Lagging Temperature": T_lag,
        "Leading Temperature": T_lead,
        "Air temperature": T_air,})

    fig = px.line(
        df_plot,
        x="Time [min]",
        y=df_plot.columns[1:],
        title=f"Cure: {cc_num} Temperature graph",
        labels={"value": "Temperature [°C]", "variable": "TC"},
    )

    fig.update_layout(hovermode="x unified")

    st.plotly_chart(fig, width='stretch')

def plot_pressure(time, Press,meetleidingen_per_zak,idx_WO,Pressure_meas_lst,Presssure_drive_lst,cc_num,AP,vac_range):
    fig = go.Figure()
    # --- Primary y-axis: Autoclave pressure ---
    fig.add_trace(
        go.Scatter(
            x=time/60,
            y=Press,
            mode="lines",
            name="AC Pressure",
            yaxis="y1",
        )
    )

    # --- Secondary y-axis: Vacuum / ML pressures ---
    meetleidingen = meetleidingen_per_zak[idx_WO]
    for idx_meet, num in enumerate(meetleidingen):
        num = num.strip("ML")
        Pressure_meas, Presssure_drive = Pressure_meas_lst[idx_WO][idx_meet], Presssure_drive_lst[idx_WO][idx_meet]
        
        fig.add_trace(
            go.Scatter(
                x=time/60,
                y=Pressure_meas,
                mode="lines",
                name=f"ML {num} Pressure",
                yaxis="y2",
            )
        )

        # --- Add green band for leaktest intervals ---
        active = np.where(Presssure_drive > 0)[0]
        if len(active) > 0:
            start_idx = np.where(np.diff(np.concatenate(([0], Presssure_drive > 0))) == 1)[0]
            end_idx   = np.where(np.diff(np.concatenate((Presssure_drive > 0, [0]))) == -1)[0]
            for s, e in zip(start_idx, end_idx):
                fig.add_vrect(
                    x0=time[s]/60,
                    x1=time[e-1]/60,
                    fillcolor="green",
                    opacity=0.3,
                    line_width=0,
                    annotation_text="Leak test" if s == start_idx[0] else "",
                    annotation_position="top left",
                )

    # --- Layout with dual y-axis ---
    fig.update_layout(
        title=f"Cure: {cc_num} Pressure graph",
        xaxis_title="Time [min]",
        yaxis=dict(
            title="Autoclave pressure [kPa]",
            range=[AP[0]-30, AP[1]+30],
            side="left",
        ),
        yaxis2=dict(
            title="Vacuum pressure [kPa]",
            overlaying="y",
            side="right",
            range=[vac_range[0]-10, vac_range[1]+10],
        ),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        hovermode="x unified",
    )

    st.plotly_chart(fig, width='stretch')

def plot_spread(time,T_lead,T_lag,cc_num):
    fig = px.line(
        x=time / 60,
        y=np.abs(T_lead - T_lag),
        labels={"x": "Time [min]", "y": "Temperature [°C]"},
        title=f"Cure: {cc_num} Temperature spread graph",
    )

    fig.add_hline(y=16, line_dash="dash", line_color="red")
    fig.update_yaxes(range=[-5, 30])
    fig.update_layout(hovermode="x unified")

    st.plotly_chart(fig, width='stretch')

def print_report(nok_found_lst,idx_WO,cc_num,CureCycle,cure_log,leaktest_log,Vacuum_log,pressure_log,spread_log,status_ok,status_nok,WO_lst):
    st.subheader("CureCycle Report")
            
    # Check if any NOK condition exists
    nok_found = nok_found_lst[idx_WO]

    # Apply conditional background color
    if nok_found:
        st.markdown('<div style="background-color:#F44336; padding:50px; border-radius:10px; color:white;">', unsafe_allow_html=True)
    else:
        st.markdown('<div style="background-color:#4CAF50; padding:15px; border-radius:10px; color:white;">', unsafe_allow_html=True)


    st.markdown(f"**CureCycle number:** {cc_num}")
    st.markdown(f'**Cure number:** {CureCycle}')
    st.markdown(f'**WO number:** {WO_lst[idx_WO]}')
    st.markdown("---")
    for log in cure_log:
        if log[0] == 0:
            st.markdown(f'{status_ok}: {log[1]}', unsafe_allow_html=True)
        if log[0] == 1:
            st.markdown(f'{status_nok}: {log[1]}', unsafe_allow_html=True)
            
    
    for log in leaktest_log:
        if log[1][0] == 0:
            st.markdown(f'{status_ok}: ML: {log[0]} {log[1][1]}', unsafe_allow_html=True)
        if log[1][0] == 1:
            st.markdown(f'{status_nok}:ML: {log[0]} {log[1][1]}', unsafe_allow_html=True)
           

    for ML_log in Vacuum_log:
        for log in ML_log[1]:
            if log[0] == 1:
                st.markdown(f'{status_nok}:  ML: {ML_log[0]} {log[1]}', unsafe_allow_html=True)
            elif log[0] == 0:
                st.markdown(f'{status_ok}: ML: {ML_log[0]} {log[1]}', unsafe_allow_html=True)

    for log in pressure_log:
        try:
            if log[0] == 0:
                st.markdown(f'{status_ok}: {log[1]}', unsafe_allow_html=True)
            if log[0] == 1:
                st.markdown(f'{status_nok}: {log[1]}', unsafe_allow_html=True)
                
        except:
            st.warning('No pressure loggings, contact PE')

    if len(spread_log)<3:
        if spread_log[0] == 0:
            st.markdown(f'{status_ok}: {spread_log[1]}', unsafe_allow_html=True)
        if spread_log[0] == 1:
            st.markdown(f'{status_nok}: {log[1]}', unsafe_allow_html=True)
          
    else:
        
        st.markdown(f'{status_nok}: Too big temperature differences, see spread.', unsafe_allow_html=True)
     

if __name__=="__main__":
#To run this file with streamlit use: python -m streamlit run Check_CureCycles/CC_checker_V2.py
#python -m streamlit run C:/Users/oscar/PycharmProjects/CureCycle_Checker/CC_checker_V2.py


    #define status colors
    status_ok = '<span style="background-color:#4CAF50; color:white; padding:3px 8px; border-radius:5px;">OK</span>'
    status_nok = '<span style="background-color:#F44336; color:white; padding:3px 8px; border-radius:5px;">NOK</span>'


    st.set_page_config(page_title="CureCycle check", layout="wide")

    # --- Title ---
    st.title("CureCycle check")
    st.markdown("Analyse test results and evaluate performance criteria.")

    # --- Input section ---
    st.sidebar.header("Curecycle number")

    with st.sidebar.form("load_form"):
        cc_num = st.text_input("Enter curecycle number", "02000007####").strip(' ')
        submitted = st.form_submit_button("Load & Analyse")

    if submitted:
        if not cc_num:
            st.warning("⚠️ Please enter a test number.")
        else:
            with st.spinner("Analysing curecycle..."):
                st.session_state.analysis = analyse_curecycle(cc_num)


    # --- Only show output if button of "load & analyse" was pressed ---
    if st.session_state.analysis:
    
        # read in  and analyse curecycle data for different WOs
        time, T_lead_lst, T_lag_lst, T_air, Press, Warning_TC_lst, cure_superlog, pressure_super_log, spread_super_log, leaktest_log, Vacuum_log, meetleidingen_per_zak, header_row,CureCycle, WO_lst, thermokoppels_per_zak, AP, vac_range, Pressure_meas_lst,  Presssure_drive_lst, nok_found_lst, is_latest = st.session_state.analysis
        
        # Show warning when the chose cure is not the latest cure.
        if not is_latest:
            st.warning('This is not the latest Cure performed. Are you sure this is the correct cure number?')

        #Provide warning if at a given point not enough TCs were recording
        for i,warning in enumerate(Warning_TC_lst):
            if warning:
                st.warning(f'Less than 3 TCs were recording at a given moment for WO: {WO_lst[i]}. Please warn PE')

        ## out dropdown menu to select different type of graphs
        graphs = ['Temperature','Pressure', 'Spread']
        selected_graph = st.sidebar.selectbox('Select a type of graph:' ,graphs)


        # --- Show WO numbers contained in this cure ---
        if "selected_WO" not in st.session_state:
            st.session_state.selected_WO = WO_lst[0]
        
        #update WOnumber for new cure check if needed.
        if st.session_state.selected_WO not in WO_lst:
            st.session_state.selected_WO = WO_lst[0]

        for WO, nok_found in zip(WO_lst, nok_found_lst):
            c1, c2 = st.sidebar.columns([0.6, 4]) #divides the width of the WO-choicemenu and the color indicators.
            # Provide a color indicator next to each WO number for clarity.
            with c1:
                st.markdown(f"<div style='height:40px;width:30px;background-color:{'#F44336' if nok_found else '#4CAF50'};'></div>",
                    unsafe_allow_html=True)
            #print the WO-number for the choice-menu
            with c2:
                if st.button(WO, key=f"WO_{WO}"):
                    st.session_state.selected_WO = WO
        
        #select the index of the chosen WO-number
        idx_WO = WO_lst.index(st.session_state.selected_WO)
        

        #select the corresponding log data of the selected WO
        cure_log = cure_superlog[idx_WO]
        pressure_log = pressure_super_log[idx_WO]
        spread_log = spread_super_log[idx_WO]
        Vacuum_log = Vacuum_log[idx_WO]
        leaktest_log = leaktest_log[idx_WO]
        T_lead = T_lead_lst[idx_WO]
        T_lag = T_lag_lst[idx_WO]
        
#==========================Now make a script to evaluate all the different results.===========================================
        # --- Layout with columns ---
        col1, col2 = st.columns([2, 1])

        # Left column: plot
        with col1:

            if selected_graph == graphs[0]:
                plot_temp(time,T_lag,T_lead,T_air)

            if selected_graph == graphs[1]:
                plot_pressure(time, Press,meetleidingen_per_zak,idx_WO,Pressure_meas_lst,Presssure_drive_lst,cc_num,AP,vac_range)

            elif selected_graph == graphs[2]:
                plot_spread(time,T_lead,T_lag,cc_num)

        # Right column: report
        with col2:
            print_report(nok_found_lst,idx_WO,cc_num,CureCycle,cure_log,leaktest_log,Vacuum_log,pressure_log,spread_log,status_ok,status_nok,WO_lst)


    else:
        st.info("Enter a test number in the sidebar and click **Load and Analyse** to begin.")


